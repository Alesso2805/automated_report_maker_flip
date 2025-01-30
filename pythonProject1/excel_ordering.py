import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

ruta_archivo = "C:/Users/Alessandro/Desktop/Ejemplo ASB.csv"
output_excel_path = "C:/Users/Alessandro/Desktop/Ejemplo_ASB_ordenado.xlsx"

# Leer el archivo CSV y procesarlo como ya lo tienes
csv_rows = []

with open(ruta_archivo, "r") as file:
    for row in file:
        csv_rows.append(row)

# Step 2: Isolate the header and data lines
header_line = csv_rows[2].strip()
data_lines = csv_rows[3:]

headers = [header.strip() for header in header_line.split(",")]

list_of_lists = []
for line in data_lines:
    info = [data.replace('"', '').strip() for data in line.split(",")]
    while len(info) > 24:
        info.pop()
    list_of_lists.append(info)

# Crear un DataFrame
df = pd.DataFrame(list_of_lists, columns=headers)

# Verificar si las columnas a ordenar existen antes de continuar
if {"Trade date", "Account Denom", "Net Amount", "Nature", "Currency", "Price", "Quantity", "Instr. Denom."}.issubset(df.columns):
    # Renombrar las columnas
    df.rename(columns={
        "Trade date": "Fecha",
        "Nature": "Operación",
        "Currency": "Moneda",
        "Price": "Px de Compra",
        "Quantity": "Cantidad",
        "Account Denom": "Cuenta Asociada",
        "Net Amount": "Importe Neto",
        "Instr. Denom.": "Ticker"
    }, inplace=True)

    cols = ["Fecha", "Operación", "Ticker", "Cantidad", "Cuenta Asociada", "Importe Neto", "Moneda", "Texto", "Px de Compra"]

    # Ordenar las columnas de Fecha de forma ascendente sin la hora
    df["Fecha"] = pd.to_datetime(df["Fecha"], format="%d/%m/%Y", errors="coerce").dt.date
    df.sort_values("Fecha", ascending=True, inplace=True)

    # Eliminar columnas duplicadas (si existieran)
    df = df.loc[:, ~df.columns.duplicated()]

    if {"Fecha", "Ticker", "Importe Neto", "Moneda"}.issubset(df.columns):
        # Reemplazar NaN con cadenas vacías y convertir a string antes de concatenar
        # Crear la columna "Texto" con el formato de fecha "DD/MM" sin el año
        df["Texto"] = df.apply(
            lambda row: f"El {row['Fecha'].strftime('%d/%m')} se recibio {row['Moneda']} {row['Importe Neto']} de {row['Ticker']} en la cuenta {row['Cuenta Asociada']}.",
            axis=1
        )

        # Insertar la columna "Texto" en la posición correcta
        df.insert(df.columns.get_loc("Px de Compra"), "Texto", df.pop("Texto"))

    # Reordenar el DataFrame
    df = df.loc[:, cols]

    # Reemplazar "Income" por "Ingresos" en la columna "Operación"
    df["Operación"] = df["Operación"].replace("Income", "Ingresos")

# Borrar columna "Op. Ccy." si existe
if "Op. Ccy." in df.columns:
    df.drop(columns=["Op. Ccy."], inplace=True)

# Guardar el DataFrame en un archivo Excel
df.to_excel(output_excel_path, index=False)

# Modificar el formato con openpyxl
wb = load_workbook(output_excel_path)
ws = wb.active

# Ajustar ancho de las columnas
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

# Centrar todo el contenido de las celdas
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Guardar los cambios en el nuevo archivo Excel
wb.save(output_excel_path)

print(f"Datos guardados exitosamente en {output_excel_path}")
